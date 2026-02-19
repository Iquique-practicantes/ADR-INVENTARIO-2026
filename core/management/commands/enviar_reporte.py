import io
from datetime import datetime
from django.core.management.base import BaseCommand
from django.core.mail import EmailMessage
from django.conf import settings
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from adr.models import (
    AllInOne, AllInOneAdmins, Notebook, MiniPC, Proyectores,
    BodegaADR, Azotea, Monitor, Audio, Tablet,
    EquiposIsla, SwitchDeRed, Televisor,
    Eliminados, HistorialCambios
)


# ── Columnas en el mismo orden de la página web ──────────────
BASE_COLUMNS = [
    ('#',                  '_row_number'),
    ('Activo',             'activo'),
    ('Estado',             'estado'),
    ('Marca',              'marca'),
    ('Modelo',             'modelo'),
    ('N° Serie',           'n_serie'),
    ('Etiqueta',           'etiqueta'),
    ('BDO',                'bdo'),
    ('NetBios',            'netbios'),
    ('Ubicación',          'ubicacion'),
    ('Registrado por',     'creado_por'),
    ('Fecha Creación',     'fecha_creacion'),
    ('Fecha Modificación', 'fecha_modificacion'),
]

COLUMNS_ASIGNADO = [
    ('#',                  '_row_number'),
    ('Activo',             'activo'),
    ('Estado',             'estado'),
    ('Marca',              'marca'),
    ('Modelo',             'modelo'),
    ('N° Serie',           'n_serie'),
    ('Etiqueta',           'etiqueta'),
    ('BDO',                'bdo'),
    ('NetBios',            'netbios'),
    ('Ubicación',          'ubicacion'),
    ('Asignado a',         'asignado_a'),
    ('Registrado por',     'creado_por'),
    ('Fecha Creación',     'fecha_creacion'),
    ('Fecha Modificación', 'fecha_modificacion'),
]

# Modelos a exportar: (clave, clase, nombre_hoja, columnas)
MODELS_CONFIG = [
    ('allinone',      AllInOne,       'All In One',              BASE_COLUMNS),
    ('allinoneadmin', AllInOneAdmins, 'All In One Admin',        BASE_COLUMNS),
    ('minipc',        MiniPC,         'Mini PC',                 BASE_COLUMNS),
    ('notebook',      Notebook,       'Notebooks',               COLUMNS_ASIGNADO),
    ('proyector',     Proyectores,    'Proyectores',             BASE_COLUMNS),
    ('bodegaadr',     BodegaADR,      'Bodega ADR',              BASE_COLUMNS),
    ('azotea',        Azotea,         'Azotea',                  BASE_COLUMNS),
    ('monitor',       Monitor,        'Monitores',               COLUMNS_ASIGNADO),
    ('audio',         Audio,          'Audio',                   BASE_COLUMNS),
    ('tablet',        Tablet,         'Tablets',                 BASE_COLUMNS),
    ('equiposisla',   EquiposIsla,    'Equipos Isla',            BASE_COLUMNS),
    ('switchdered',   SwitchDeRed,    'Switch De Red',           BASE_COLUMNS),
    ('televisor',     Televisor,      'Televisores',             BASE_COLUMNS),
]


def _resolve_cell(obj, field_name, row_idx):
    """Resuelve el valor de una celda según el tipo de campo."""
    if field_name == '_row_number':
        return row_idx

    if field_name in ('creado_por', 'eliminado_por', 'usuario'):
        user = getattr(obj, field_name, None)
        if user is None:
            return 'Admin' if field_name == 'creado_por' else ''
        name = user.get_full_name()
        return name if name.strip() else user.username

    if field_name in ('fecha_creacion', 'fecha_modificacion', 'fecha_eliminacion'):
        val = getattr(obj, field_name, None)
        return val.strftime('%d/%m/%Y') if val else ''

    val = getattr(obj, field_name, None)
    return str(val) if val is not None else ''


def _fill_sheet(ws, queryset, columns):
    """Rellena una hoja Excel con encabezados y datos."""
    # Encabezados
    for col_num, (header, _) in enumerate(columns, 1):
        ws[f"{get_column_letter(col_num)}1"] = header

    # Datos
    for row_num, obj in enumerate(queryset, 2):
        row_idx = row_num - 1
        for col_num, (_, field_name) in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            try:
                ws[f"{col_letter}{row_num}"] = _resolve_cell(obj, field_name, row_idx)
            except Exception:
                ws[f"{col_letter}{row_num}"] = ''


class Command(BaseCommand):
    help = 'Genera un reporte Excel de todos los activos y lo envía por correo'

    def handle(self, *args, **options):
        self.stdout.write(self.style.SUCCESS('Iniciando generación de reporte...'))

        # Generar Excel en memoria con openpyxl
        wb = Workbook()
        # Eliminar la hoja por defecto
        wb.remove(wb.active)

        models_processed = 0

        for key, model_class, sheet_name, columns in MODELS_CONFIG:
            ws = wb.create_sheet(title=sheet_name[:31])

            # select_related para evitar N+1 queries
            fk_fields = []
            for f in model_class._meta.fields:
                if f.is_relation and f.name in ('creado_por', 'eliminado_por', 'usuario'):
                    fk_fields.append(f.name)

            queryset = model_class.objects.all()
            if fk_fields:
                queryset = queryset.select_related(*fk_fields)

            _fill_sheet(ws, queryset, columns)
            models_processed += 1
            self.stdout.write(f'Agregado hoja: {sheet_name} ({queryset.count()} registros)')

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Enviar correo
        date_str = datetime.now().strftime('%Y-%m-%d')
        filename = f'Inventario_ADR_{date_str}.xlsx'

        subject = f'Reporte de Inventario ADR - {date_str}'
        body = f'Adjunto encontrará el reporte actualizado del inventario con {models_processed} categorías de equipos.'
        to_emails = settings.EMAIL_RECIPIENTS

        # DEBUG: Mostrar lista completa de destinatarios
        self.stdout.write(self.style.WARNING(f'DEBUG - Lista de destinatarios configurada: {to_emails}'))
        self.stdout.write(self.style.WARNING(f'DEBUG - Total de destinatarios: {len(to_emails)}'))

        # Crear y enviar el correo
        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=to_emails
        )

        email.attach(filename, output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        try:
            email.send()
            self.stdout.write(self.style.SUCCESS(f'Reporte enviado exitosamente a {len(to_emails)} destinatarios:'))
            for destinatario in to_emails:
                self.stdout.write(self.style.SUCCESS(f'  - {destinatario}'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error al enviar el reporte: {str(e)}'))
