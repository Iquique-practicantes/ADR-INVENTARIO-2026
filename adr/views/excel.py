"""
Vistas para Descarga de Datos en Formato Excel
"""
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import redirect
from django.utils import timezone
from django.views.generic import TemplateView
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from adr.decorators import add_group_name_to_context
from adr.models import (
    AllInOne, AllInOneAdmins, Notebook, MiniPC, Proyectores,
    BodegaADR, Azotea, Monitor, Audio, Tablet,
    EquiposIsla, SwitchDeRed, Televisor,
    Eliminados, HistorialCambios
)

# ── Definición de columnas (header, field_name) ──────────────
# Orden idéntico al de las tablas en la web
BASE_COLUMNS = [
    ('#',                  '_row_number'),
    ('Activo',             'activo'),
    ('Estado',             'estado'),
    ('Marca',              'marca'),
    ('Modelo',             'modelo'),
    ('N° Serie',           'n_serie'),
    ('Etiqueta',           'etiqueta'),
    ('BDO',                'bdo'),
    ('NetBios',            'netbios'),
    ('Ubicación',          'ubicacion'),
    ('Registrado por',     'creado_por'),
    ('Fecha Creación',     'fecha_creacion'),
    ('Fecha Modificación', 'fecha_modificacion'),
]

# Modelos con campo asignado_a (Notebook, Monitor)
COLUMNS_ASIGNADO = [
    ('#',                  '_row_number'),
    ('Activo',             'activo'),
    ('Estado',             'estado'),
    ('Marca',              'marca'),
    ('Modelo',             'modelo'),
    ('N° Serie',           'n_serie'),
    ('Etiqueta',           'etiqueta'),
    ('BDO',                'bdo'),
    ('NetBios',            'netbios'),
    ('Ubicación',          'ubicacion'),
    ('Asignado a',         'asignado_a'),
    ('Registrado por',     'creado_por'),
    ('Fecha Creación',     'fecha_creacion'),
    ('Fecha Modificación', 'fecha_modificacion'),
]

# Mapeo modelo → columnas personalizadas
COLUMN_MAP = {
    'notebook': COLUMNS_ASIGNADO,
    'monitor':  COLUMNS_ASIGNADO,
}

# Modelos que usan la exportación genérica (todos los campos del modelo)
GENERIC_MODELS = {'eliminados', 'historialcambios'}


def _resolve_cell(obj, field_name, row_idx):
    """Resuelve el valor de una celda según el tipo de campo."""
    if field_name == '_row_number':
        return row_idx

    if field_name in ('creado_por', 'eliminado_por', 'usuario'):
        user = getattr(obj, field_name, None)
        if user is None:
            return 'Admin' if field_name == 'creado_por' else ''
        name = user.get_full_name()
        return name if name.strip() else user.username

    if field_name in ('fecha_creacion', 'fecha_modificacion', 'fecha_eliminacion'):
        val = getattr(obj, field_name, None)
        return val.strftime('%d/%m/%Y') if val else ''

    val = getattr(obj, field_name, None)
    return str(val) if val is not None else ''


@add_group_name_to_context
class DescargarExcelView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista para descargar datos en formato Excel"""

    def test_func(self):
        """Solo ADR o Operadores ADR pueden descargar"""
        return self.request.user.groups.filter(
            name__in=['ADR', 'Operadores ADR', 'Auxiliares Operadores ADR']
        ).exists()

    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def get(self, request, *args, **kwargs):
        """Genera y descarga archivo Excel del modelo especificado"""
        model_name = kwargs.get('model_name')
        fecha_actual = timezone.now().strftime('%d-%m-%Y')

        # Mapeo de modelos a clases y nombres de archivo
        model_mapping = {
            'allinone': (AllInOne, f'AllInOne_{fecha_actual}.xlsx'),
            'allinoneadmin': (AllInOneAdmins, f'AllInOneAdmins_{fecha_actual}.xlsx'),
            'equiposisla': (EquiposIsla, f'EquiposIsla_{fecha_actual}.xlsx'),
            'notebook': (Notebook, f'Notebooks_{fecha_actual}.xlsx'),
            'minipc': (MiniPC, f'MiniPCs_{fecha_actual}.xlsx'),
            'proyector': (Proyectores, f'Proyectores_{fecha_actual}.xlsx'),
            'bodegaadr': (BodegaADR, f'BodegaADR_{fecha_actual}.xlsx'),
            'azotea': (Azotea, f'Azotea_{fecha_actual}.xlsx'),
            'eliminados': (Eliminados, f'Eliminados_{fecha_actual}.xlsx'),
            'historialcambios': (HistorialCambios, f'HistorialCambios_{fecha_actual}.xlsx'),
            'monitor': (Monitor, f'Monitores_{fecha_actual}.xlsx'),
            'audio': (Audio, f'Audio_{fecha_actual}.xlsx'),
            'tablet': (Tablet, f'Tablets_{fecha_actual}.xlsx'),
            'switchdered': (SwitchDeRed, f'SwitchDeRed_{fecha_actual}.xlsx'),
            'televisor': (Televisor, f'Televisores_{fecha_actual}.xlsx'),
        }

        # Validación del modelo
        if model_name not in model_mapping:
            return HttpResponse(status=404)

        model_class, filename = model_mapping[model_name]

        # Aplicar filtro de búsqueda si existe
        search_query = request.GET.get('search', '').strip()
        if search_query:
            # Construir filtro dinámicamente según los campos disponibles
            q_filters = Q()
            search_fields = [
                'activo', 'marca', 'modelo', 'n_serie',
                'etiqueta', 'ubicacion',
            ]
            for sf in search_fields:
                if any(f.name == sf for f in model_class._meta.fields):
                    q_filters |= Q(**{f'{sf}__icontains': search_query})
            # Campos opcionales
            if any(f.name == 'netbios' for f in model_class._meta.fields):
                q_filters |= Q(netbios__icontains=search_query)
            if any(f.name == 'bdo' for f in model_class._meta.fields):
                q_filters |= Q(bdo__icontains=search_query)
            # Búsqueda por nombre de usuario creador
            if any(f.name == 'creado_por' for f in model_class._meta.fields):
                q_filters |= (
                    Q(creado_por__first_name__icontains=search_query) |
                    Q(creado_por__last_name__icontains=search_query)
                )
            queryset = model_class.objects.filter(q_filters)
        else:
            queryset = model_class.objects.all()

        # select_related para evitar N+1 queries en ForeignKeys
        fk_fields = []
        for f in model_class._meta.fields:
            if f.is_relation and f.name in ('creado_por', 'eliminado_por', 'usuario'):
                fk_fields.append(f.name)
        if fk_fields:
            queryset = queryset.select_related(*fk_fields)

        # Elegir columnas
        if model_name in GENERIC_MODELS:
            columns_raw = [field.name for field in model_class._meta.fields]
            columns = [(col.replace('_', ' ').title(), col) for col in columns_raw]
        else:
            columns = COLUMN_MAP.get(model_name, BASE_COLUMNS)

        # Crear archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = model_name.capitalize()

        # Encabezados
        for col_num, (header, _) in enumerate(columns, 1):
            ws[f"{get_column_letter(col_num)}1"] = header

        # Datos
        for row_num, obj in enumerate(queryset, 2):
            row_idx = row_num - 1  # Número secuencial empezando en 1
            for col_num, (_, field_name) in enumerate(columns, 1):
                col_letter = get_column_letter(col_num)
                try:
                    ws[f"{col_letter}{row_num}"] = _resolve_cell(obj, field_name, row_idx)
                except Exception:
                    ws[f"{col_letter}{row_num}"] = ''

        # Configurar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response
